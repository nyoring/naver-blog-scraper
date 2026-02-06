"""
Flask 웹 애플리케이션 - 네이버 블로그 스크래퍼
"""

import json
import io
import threading
from flask import Flask, render_template, request, jsonify, Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from scraper import count_posts, scrape_all_posts

app = Flask(__name__)

# 필드 정의 (순서, 한글 라벨, Excel 열 너비)
FIELD_ORDER = [
    "title",
    "url",
    "content",
    "author",
    "blog_name",
    "date",
    "likes",
    "comments",
]
FIELD_LABELS = {
    "title": "제목",
    "url": "URL",
    "content": "내용",
    "author": "작성자",
    "blog_name": "블로그명",
    "date": "날짜",
    "likes": "좋아요",
    "comments": "댓글",
}
FIELD_WIDTHS = {
    "title": 40,
    "url": 50,
    "content": 80,
    "author": 15,
    "blog_name": 20,
    "date": 15,
    "likes": 10,
    "comments": 10,
}

# 글로벌 스크래핑 세션 상태 (단일 세션 모델)
scrape_session = {
    "active": False,
    "pause_event": None,  # threading.Event: set=실행중, clear=일시정지
    "stop_event": None,  # threading.Event: set=정지요청
}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/count", methods=["POST"])
def api_count():
    """게시물 수 조회 API"""
    data = request.get_json()
    keyword = data.get("keyword", "").strip()
    start_date = data.get("startDate", "").strip()
    end_date = data.get("endDate", "").strip()

    if not keyword or not start_date or not end_date:
        return jsonify({"error": "키워드, 시작일, 종료일을 모두 입력해주세요."}), 400

    try:
        total = count_posts(keyword, start_date, end_date)
        return jsonify({"totalCount": total, "keyword": keyword})
    except Exception as e:
        return jsonify({"error": f"조회 중 오류가 발생했습니다: {str(e)}"}), 500


@app.route("/api/scrape")
def api_scrape():
    """게시물 스크래핑 API (Server-Sent Events)"""
    keyword = request.args.get("keyword", "").strip()
    start_date = request.args.get("startDate", "").strip()
    end_date = request.args.get("endDate", "").strip()
    fields_param = request.args.get("fields", "").strip()
    content_mode = request.args.get("contentMode", "preview").strip()

    if not keyword or not start_date or not end_date:
        return jsonify({"error": "파라미터 부족"}), 400

    # 필드 파싱: 쉼표 구분 문자열 → 리스트, 빈 값이면 None (전체 필드)
    fields = (
        [f.strip() for f in fields_param.split(",") if f.strip()]
        if fields_param
        else None
    )

    # content_mode 검증
    if content_mode not in ("preview", "full", "none"):
        content_mode = "preview"

    # 활성 필드 목록 결정 (SSE 이벤트 데이터에 포함할 키)
    if fields is None:
        active_fields = list(FIELD_ORDER)
    else:
        active_fields = ["title"]  # title은 항상 포함
        for f in FIELD_ORDER:
            if f != "title" and f in fields:
                active_fields.append(f)
        if content_mode != "none" and "content" not in active_fields:
            active_fields.append("content")

    def generate():
        # 기존 세션이 있으면 강제 정지
        if scrape_session["active"] and scrape_session["stop_event"]:
            scrape_session["stop_event"].set()
            # 일시정지 상태면 풀어서 제너레이터가 정지 확인할 수 있게
            if scrape_session["pause_event"]:
                scrape_session["pause_event"].set()

        # 새 이벤트 인스턴스 생성
        pause_event = threading.Event()
        pause_event.set()  # set=실행중 (clear=일시정지)
        stop_event = threading.Event()  # 초기: 정지 안 됨

        scrape_session["active"] = True
        scrape_session["pause_event"] = pause_event
        scrape_session["stop_event"] = stop_event

        try:
            total = count_posts(keyword, start_date, end_date)
            yield f"event: init\ndata: {json.dumps({'total': total})}\n\n"

            for detail in scrape_all_posts(
                keyword,
                start_date,
                end_date,
                progress_callback=None,
                fields=fields,
                content_mode=content_mode,
                pause_event=pause_event,
                stop_event=stop_event,
            ):
                event_data = {}
                for f in active_fields:
                    if f == "content" and content_mode == "none":
                        continue
                    if f == "content" and content_mode == "preview":
                        # snippet을 content로 매핑
                        event_data["content"] = detail.get(
                            "content", detail.get("snippet", "")
                        )
                    else:
                        event_data[f] = detail.get(
                            f, "" if f not in ("likes", "comments") else 0
                        )
                yield f"data: {json.dumps(event_data, ensure_ascii=False)}\n\n"

            # 제너레이터 종료 이유 판별
            if stop_event.is_set():
                yield "event: stopped\ndata: {}\n\n"
            else:
                yield "event: done\ndata: {}\n\n"

        except Exception as e:
            error_data = {"error": str(e)}
            yield f"event: error\ndata: {json.dumps(error_data, ensure_ascii=False)}\n\n"
        finally:
            scrape_session["active"] = False
            scrape_session["pause_event"] = None
            scrape_session["stop_event"] = None

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.route("/api/pause", methods=["POST"])
def api_pause():
    """스크래핑 일시정지 API"""
    if not scrape_session["active"] or scrape_session["pause_event"] is None:
        return jsonify({"error": "활성 스크래핑 세션이 없습니다"}), 409
    scrape_session["pause_event"].clear()  # clear = 일시정지 (wait()에서 블록)
    return jsonify({"status": "paused"})


@app.route("/api/resume", methods=["POST"])
def api_resume():
    """스크래핑 재개 API"""
    if not scrape_session["active"] or scrape_session["pause_event"] is None:
        return jsonify({"error": "활성 스크래핑 세션이 없습니다"}), 409
    scrape_session["pause_event"].set()  # set = 실행중 (wait()에서 즉시 반환)
    return jsonify({"status": "resumed"})


@app.route("/api/stop", methods=["POST"])
def api_stop():
    """스크래핑 정지 API"""
    if not scrape_session["active"] or scrape_session["stop_event"] is None:
        return jsonify({"error": "활성 스크래핑 세션이 없습니다"}), 409
    scrape_session["stop_event"].set()  # 정지 시그널
    # 일시정지 상태면 풀어서 제너레이터가 정지를 확인할 수 있게 함
    if scrape_session["pause_event"]:
        scrape_session["pause_event"].set()
    return jsonify({"status": "stopped"})


@app.route("/api/export-excel", methods=["POST"])
def api_export_excel():
    """스크래핑 결과를 Excel 파일로 다운로드"""
    data = request.get_json()
    results = data.get("results", [])
    fields_param = data.get("fields", [])
    content_mode = data.get("contentMode", "preview")

    # 활성 필드 결정 (canonical order 유지)
    if not fields_param:
        active_fields = list(FIELD_ORDER)
    else:
        active_fields = ["title"]
        for f in FIELD_ORDER:
            if f == "title":
                continue
            if f == "content":
                if content_mode != "none":
                    active_fields.append("content")
                continue
            if f in fields_param:
                active_fields.append(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "블로그 스크래핑 결과"

    # 헤더 스타일
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2E7D32", end_color="2E7D32", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더 작성 (동적)
    for col, field in enumerate(active_fields, 1):
        cell = ws.cell(row=1, column=col, value=FIELD_LABELS.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 스타일
    data_font = Font(name="맑은 고딕", size=10)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    # 데이터 작성 (동적)
    for row_idx, item in enumerate(results, 2):
        for col, field in enumerate(active_fields, 1):
            value = item.get(field, "")
            if field == "content" and isinstance(value, str):
                value = value[:32000]  # Excel 셀 최대 글자수 제한
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 열 너비 설정 (동적)
    for i, field in enumerate(active_fields, 1):
        ws.column_dimensions[chr(64 + i)].width = FIELD_WIDTHS.get(field, 15)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 파일 전송
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="blog_scraping_result.xlsx",
    )


# Chromium 설치 상태 (app_entry.py에서 업데이트)
chromium_install_state = {"status": "idle", "percent": 0, "message": ""}


@app.route("/setup")
def setup_page():
    return render_template("setup.html")


@app.route("/api/setup-status")
def api_setup_status():
    def generate():
        import time

        while chromium_install_state["status"] in ("idle", "downloading"):
            data = json.dumps(
                {
                    "percent": chromium_install_state["percent"],
                    "message": chromium_install_state["message"],
                },
                ensure_ascii=False,
            )
            yield f"data: {data}\n\n"
            time.sleep(0.5)
        if chromium_install_state["status"] == "done":
            yield "event: done\ndata: {}\n\n"
        elif chromium_install_state["status"] == "error":
            error_data = json.dumps(
                {"error": chromium_install_state["message"]}, ensure_ascii=False
            )
            yield f"event: error\ndata: {error_data}\n\n"

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


if __name__ == "__main__":
    app.run(debug=True, port=8000, threaded=True)
